import datetime
import logging
import os
import sys
from configparser import ConfigParser
from pathlib import Path

import psycopg2
from openpyxl import load_workbook


class LoggerServices:
    """
    This class contains the various methods needed to log and notify
    messages in a log file for each script run
    
     Args
        Required: none
        Optional: none

    Alerts: Critical | WARN | ERROR

    Logging: none
    
    """

    __author__ = "Barry Onizak"
    __version__ = "20220403.3"
    # # # # # End of header # # # #

    log_file = ""
    logLevel = ""
    AppConfig_AOD = None

    def getLogger(self, name):
        """
        This method creates and returns a object used to log each script run
        """
        log_file = self.setLogFile()
        self.logLevel = self.setLogLevel()
        logger = logging.getLogger(name)
        logging.basicConfig(filename=log_file, level=self.getLogLevel(),
                            format=' %(asctime)s %(levelname)s: %(message)s', datefmt='%Y/%m/%d %H:%M:%S')

        return logger

    @staticmethod
    def date():
        """
        This method returns a formatted string of the date in YYYY/MM/DD/hh/mm/ss format
        """
        return datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')

    @staticmethod
    def file_date():
        """
        This method returns a formatted date string in YYYYMMDD for appending to file names
        """
        return datetime.datetime.now().strftime('%Y%m%d')

    def getLogDir(self):
        """
        This method sets the logs directory
        """
        log_dir = os.path.join(str(Path.home()), 'logs', self.getScriptName())

        if os.path.isdir(log_dir):
            return log_dir
        else:
            os.makedirs(log_dir)
            if not os.path.exists(log_dir):
                msg = 'Initial setup of logs dir not done, exiting script run!'
                print(f'{self.date()}: {msg}')
                sys.exit(1)
            else:
                return log_dir

    def setLogFile(self):
        """
        This method sets log file name to be written to for each script run
        """
        return os.path.join(self.getLogDir(),
                            f'{str(os.path.basename(sys.argv[0])).replace(".py", "")}.{self.file_date()}.log')

    def starting_template(self, parameter_list, args):
        """
        This method takes a list of cmd line args
        passed and displays each running script in a similar view
        """
        self.info(f'Starting {self.getScriptName()}')
        self.info(f"Extracting input params: {(' '.join(map(str, parameter_list)))}")
        self.info(f"Log Level {self.getLogLevel()}")
        self.info(f"Script Version {self.get_script_version(args)}")
        return None

    def setLogLevel(self):
        resource_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "resource")
        log_level = "DEBUG"
        dbConn = None
        try:
            # read connection parameters
            params = self.dbParser(os.path.join(resource_path, 'database.ini'), 'postgresql')
            dbConn = psycopg2.connect(**params)
        except (Exception, psycopg2.DatabaseError) as error:
            print(f'Error connecting to PostgreSQL {error}')

        try:
            cur = dbConn.cursor()
            sql = "SELECT id, key, value FROM \"public\".\"AppConfig\" WHERE key = 'Log_Level'"

            cur.execute(sql)
            if cur.rowcount == 1:
                rowData = cur.fetchone()
                log_level = str(rowData[2]).upper()
            else:
                print("Invalid rows from AppConfig table")
            cur.close()

        except (Exception, psycopg2.DatabaseError) as error:
            print(f'An ERROR occurred when querying AppConfig: {error}. Setting LOG LEVEL to DEBUG!.')

            log_level = "DEBUG"
        finally:
            dbConn.close()
            return log_level

    def getLogLevel(self):
        return self.logLevel

    @staticmethod
    def get_script_version(args):
        if hasattr(args, 'version'):
            return args.version
        else:
            return "UNKNOWN"

    def openLogFile(self):
        """
        This method opens the logfile and prepends the starting info
        """
        self.log_file = self.setLogFile()
        try:
            if os.path.exists(self.log_file):
                fo = open(self.log_file, 'a', encoding='utf-8')
                print(f'\nUsing logfile {self.log_file}')
            else:
                fo = open(self.log_file, 'w', encoding='utf-8')
                print(f'\nCreated logfile {self.log_file}')

            fo.write("\n\n")
            fo.write(self.startScriptLine())
            fo.write("\n")
            fo.close()
        except IOError as ioe:
            msg = f'Log file {self.log_file} write error. {ioe}.'
            print(f'{self.date()}: {msg}')

        return self.log_file

    def closeLogFile(self):
        """
        This method appends the closing log info and closes the logfile
        """
        self.log_file = self.setLogFile()
        try:
            fo = open(self.log_file, 'a', encoding='utf-8')
            fo.write("\n")
            fo.write(self.endingScriptLine())
            fo.write("\n")
            fo.close()
        except IOError as ioe:
            msg = f'Log file {self.log_file} write error. {ioe}.'
            print(f'{self.date()}: {msg}')

        return self.log_file

    @staticmethod
    def getScriptName():
        """
        This method returns the script name
        """
        return str(os.path.basename(sys.argv[0])).split('.')[0]

    def startScriptLine(self):
        """
        This method returns the script logfile opening info
        """
        script_path = os.path.normpath(os.path.join(os.popen("pwd").read().strip('\n'), str(sys.argv[0])))
        if not os.path.isfile(script_path):
            print('No such script name in toolkit folder..exiting')
            sys.exit(1)

        script_start = f'{self.separationBar()} \n Starting script {script_path} \n{self.separationBar()}'
        return script_start

    def endingScriptLine(self):
        """
        This method takes a list of cmd line args
        passed and displays each running script in a similar view
        """
        script_path = os.path.normpath(os.path.join(os.popen("pwd").read().strip('\n'), str(sys.argv[0])))
        script_end = f'{self.separationBar()} \n End of script {script_path} \n{self.separationBar()}'
        return script_end

    def set_log_level(self):
        wb = load_workbook(os.path.join(self.resource_path, "BackupList.xlsx"))
        sheetset = {'AppConfig': 6}

        for ws in wb:
            if ws.title in sheetset.keys():
                worksheet = wb[ws.title]

                row_sets = [worksheetsets for worksheetsets in worksheet.iter_rows(
                    min_row=2, max_col=sheetset[ws.title], min_col=1, values_only=True)]

                self.log_level = row_sets[2][1]
        return self.log_level

    def dbParser(self, filename, section='postgresql'):
        # create a parser
        parser = ConfigParser()
        # read config file
        parser.read(filename)

        # get section, default to postgresql
        db = {}
        if parser.has_section(section):
            params = parser.items(section)
            for param in params:
                db[param[0]] = param[1]
        else:
            raise Exception('A Section {0} not found in the {1} file'.format(section, filename))

        return db

    @staticmethod
    def separationBar():
        """
        This method returns a 30 char separation bar
        """
        return 75 * f'='

    def critical(self, msg):
        """
        This method takes a message and logs it as a
        CRITICAL to the script run log
        """
        logging.critical(msg)
        return None

    def error(self, msg):
        """
        This method takes a message and logs it as an
        ERROR  to the script run log
        """
        logging.error(msg)
        return None

    def warn(self, msg):
        """
        This method takes a message and logs it as an
        WARNING to the script run log
        """
        logging.warning(msg)
        return None

    def info(self, msg):
        """
        This method takes a message and logs it as an
        INFO to the script run log
        """
        logging.info(msg)
        return None

    def debug(self, msg):
        """
        This method takes a message and logs it as a
        DEBUG message to the script run log
        """
        logging.debug(msg)
        return None
